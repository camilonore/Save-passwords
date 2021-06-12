from openpyxl import load_workbook
import pandas as pd
import random
import argparse

#########################__Crypt_and_decrypt__##################################


__author__ = "Shangru Li"
__copyright__ = "Copyright 2021, Shangru Li"
__credits__ = "Shangru Li"
__license__ = "MIT"
__version__ = "3.4"
__maintainer__ = "Shangru Li"
__email__ = "maxsli@protonmail.com"
__status__ = "Stable"

###############################___Seeds___######################################
seed_letter: list = ["a", "b", "c", "d", "e", "f", "g", "h", "i", "j", "k", "l", "m", "n", "o", "p",
                     "q", "r", "s", "t", "u", "v", "w", "x", "y", "z", ]
seed_number: list = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
seed_symbol: list = ["`", "~", "!", "@", "#", "$", "%", "^", "&", "*", "(", ")", "-", "_", "+", "=",
                     "[", "{", "]", "}", "|", ";", ":", ">", ",", "<", ".", "/", "?", ]
###############################___Args___#######################################
parser = argparse.ArgumentParser()
parser.add_argument("-d", "--decrypt",
                    help="Decrypt an encrypted string", type=str)
parser.add_argument("-e", "--encrypt",
                    help="Encrypt an input string", type=str)


def generate_seed_indicator(seed_symbol: list) -> str:
    """
    Generate a special character to separate the seed and code from the `seed_symbol`.
    """
    seed_indicator: str = seed_symbol[random.randint(0, len(seed_symbol) - 1)]
    seed_symbol.remove(seed_indicator)
    return seed_indicator


def get_random_case_seedletter() -> str:
    """
    Randomly pick a either uppercase or lowercase letter from `seedLetter`.
    """
    if random.randint(0, 1) == 0:
        return seed_letter[random.randint(0, len(seed_letter) - 1)]
    else:
        return str.capitalize(seed_letter[random.randint(0, len(seed_letter) - 1)])


def encrypt(text_to_encrypt: str, is_seed: bool = False, seed_symbol_copy: list = []) -> str:
    """
    Encrypt the given string `text_to_encrypt` and return the result.
    """
    # Pre-condition: Input `text_to_encrypt` should not be empty
    if text_to_encrypt is None or text_to_encrypt == '':
        raise SyntaxError("Input cannot be empty.")
    if is_seed == False:
        # Make a copy of the current `seed_symbol`
        seed_symbol_copy = seed_symbol.copy()
        # Pick a random seedIndicator
        seed_indicator: str = generate_seed_indicator(seed_symbol_copy)
    else:
        seed_indicator = text_to_encrypt[0]
        # Remove the `seedIndicator`
        text_to_encrypt = text_to_encrypt.replace(seed_indicator, "")
    code: str = ""
    # Initialize a random offset to be added to the character's ASCII code
    offset: int = random.randint(10, 99)
    # The first two characters of the `seed` is the `offset`
    seed: str = str(offset)
    for n in text_to_encrypt:
        # Randomly pick a letter from 'seed_letter', upper or lower case
        random_letter: str = get_random_case_seedletter()
        # Randomly pick a number from 'seed_number'
        random_number: str = seed_number[random.randint(
            0, len(seed_number) - 1)]
        # Randomly pick a symbol from 'seed_symbol'
        random_symbol: str = seed_symbol_copy[random.randint(
            0, len(seed_symbol_copy) - 1)]
        # Appending to the seed
        seed_combine: str = random_letter + random_number + random_symbol
        # Appending to the seed
        seed = seed + seed_combine
        # The code is:
        # ASCII code of character plus 'offset' + combination of three seeds
        code = code + str(ord(n) + offset) + seed_combine
    if is_seed == True:
        # We have encrypted the seed. Return the full encrypted input.
        return seed + seed_indicator + code
    else:
        # Else we encode the seed
        return (
            encrypt(seed_indicator + seed, True, seed_symbol_copy)
            + seed_indicator
            + code
            + seed_indicator
        )


def decrypt(text_to_decrypt: str, is_seed: bool = False) -> str:
    """
    Decrypt the given cypher `text_to_decrypt` and return the result.
    """
    try:
        # The last character is the `seedIndicator`
        seed_indicator: str = text_to_decrypt[len(text_to_decrypt) - 1]
        # Removing the `seedIndicator` from input
        text_to_decrypt = text_to_decrypt[:-1]
        # The `textToDecrpt` should looks like this:
        # {seedForEncryptedSeed + seedIndicator + encryptedSeed} +
        # {seedIndicator + encryptedInput + seedIndicator}
        if is_seed == False:
            # Splitting the seed and input
            (
                seed_for_encrypted_seed,
                encrypted_seed,
                encrypted_input,
            ) = text_to_decrypt.split(seed_indicator)
            # Decrpt the seed for user input
            decrypted_seed: str = decrypt(
                seed_for_encrypted_seed
                + seed_indicator
                + encrypted_seed
                + seed_indicator,
                True,
            )
            # Combine the decrypted seed with encrypted input for further encrypt
            text_to_decrypt = decrypted_seed + seed_indicator + encrypted_input
        # The first three characters of seed is the `offset`
        offset: int = int(text_to_decrypt[0] + text_to_decrypt[1])
        # Remove the `offset` from `textToDecrpt`
        text_to_decrypt = text_to_decrypt[2: len(text_to_decrypt)]
        # Split the textToDecrypt by seedIndicator to get the seed and code
        seed, code = text_to_decrypt.split(seed_indicator)
        result: list = []
        # This is to remember the ASCII code for a paticular character
        # For example, ASCII code 102, in this list: ['1', '0', '2']
        ascii_list: list = []
        # Check which seed the index is pointing at
        check: int = 0
        # The index for seed list
        seed_index: int = 0
        for i in range(len(code)):
            # separate the three seeds
            seed1: str = seed[seed_index]
            seed2: str = seed[seed_index + 1]
            seed3: str = seed[seed_index + 2]
            # current index points to `seed1`, increment `check` go to next character
            if code[i] == seed1 and check == 0:
                check = 1
            # current index points to `seed2`, do the same
            elif code[i] == seed2 and check == 1:
                check = 2
            # current index points to `seed3`, we can start decrypting
            elif code[i] == seed3 and check == 2:
                # each original character is the ASCII code - offset
                result.append((chr(int("".join(ascii_list)) - offset)))
                # clear the array after the have decrypted
                ascii_list.clear()
                # Go to the next set of seed
                seed_index += 3
                check = 0
            else:
                # index points to the encrypted char, store it for decrypt
                ascii_list.append(code[i])
        return "".join(result)
    except:
        raise SyntaxError(
            "Decryption failed, please make sure the encrypted text is correct.")

###########################################################################


book = 'Python\Tkinter\Password.xlsx'


class Functions:
    def __init__(self, function=None, domain='None', email='None', password='None', newpassword=None, privatekey=None) -> None:
        self.newpassword = newpassword
        self.privateKey = privatekey
        self.function = function
        self.domain = domain
        self.email = email
        self.password = password
        # Cargar el libro .xlsx
        self.wb = load_workbook(book)
        self.ws = self.wb.active
        self.ws.tittle = 'Passwords'
        # Leer el contador asignado en el libro .xlsx
        self.counter = self.ws['F1'].value
        # Seleccion del metodo que se va a usar
        if self.function == 'add':
            self.AddData()
        elif self.function == 'delete':
            self.DeleteData()
        elif self.function == 'change':
            self.ChangePass(self.domain, self.newpassword)
        elif self.function == 'privatekey':
            self.privateKeyFunct()

    # Metodo comprobacion de asignacion de la llave privada
    def checkprivatetime(self):
        if self.ws['F1'].value == 2:
            self.counter += 1
            # Asign the new counter value
            self.ws['F1'].value = self.counter
            # Save the book
            self.wb.save(book)
            return True

    # Metodo llave privada
    def privateKeyFunct(self):
        # Guardar el valor de la llave privada
        self.ws['F2'].value = encrypt(self.privateKey)
        # Guardar el libro
        self.wb.save(book)

    # Funcionalidad para añadir una contraseña
    def AddData(self):
        # Asign values
        self.ws['A'+str(self.counter)].value = self.domain
        self.ws['B'+str(self.counter)].value = encrypt(self.email)
        self.ws['C'+str(self.counter)].value = encrypt(self.password)
        # Increasing counter value
        self.counter += 1
        # Asign the new counter value
        self.ws['F1'].value = self.counter
        # Save the book
        self.wb.save(book)

    # Funcionalidad para borra contraseña
    def DeleteData(self):
        # Ejecutar el metodo LookFor para buscar la posicion del dominio
        DomainPositioin = self.LookFor(self.domain, self.counter)
        # Borrando los datos
        self.ws.delete_rows(DomainPositioin)
        # Save the book
        self.wb.save(book)

    # Funcionalidad para conocer una contraseña
    def KnowPassword(self, enckey):
        # Ejecutar el metodo LookFor para buscar la posicion de la contraseña
        PasswordPosition = 'C'+str(self.LookFor(self.domain, self.counter))
        # Encontrar el valor de la contraseña
        PasswordValue = self.ws[PasswordPosition].value
        # Verificacion de la contraseña
        # comprbar la llave privada
        if enckey == decrypt(self.ws['F2'].value):
            PasswordValue = decrypt(PasswordValue)
            return PasswordValue
        else:
            return 'Wrong private key!!'

    # Funcionalidad para cambiar una contraseña
    def ChangePass(self, domain, newpassword):
        # Ejecutar el metodo LookFor para buscar la posicion de la contraseña
        PasswordPosition = 'C'+str(self.LookFor(domain, self.counter))
        # Asignando el nuevo valor a la contraseña
        self.ws[PasswordPosition] = encrypt(newpassword)
        # Save the book
        self.wb.save(book)

    # Funcionalidad para buscar contraseñas guardadas en el archivo xlsx
    def LookFor(self, domain, counter):
        # Determinar un limite de busqueda en el libro con counter
        for i in range(0, counter):
            df = pd.read_excel(
                book, "Sheet")
            # Buscar si el nombre en la posicion es igual al dominio ingresado
            if df.iloc[i, 0] == domain:
                # Devolver la posicion en la que el dominio se encuentra
                return (i+2)


if __name__ == '__main__':
    Functions
